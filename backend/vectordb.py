from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct
from sentence_transformers import SentenceTransformer
import uuid
import os

class VectorDB:
    def __init__(self):
        self.client = QdrantClient(
            host=os.getenv("QDRANT_HOST", "localhost"),
            port=int(os.getenv("QDRANT_PORT", 6333))
        )
        self.collection_name = "rrhh_docs"
        self.encoder = SentenceTransformer('paraphrase-multilingual-MiniLM-L12-v2')
        self._init_collection()
    
    def _init_collection(self):
        try:
            self.client.get_collection(self.collection_name)
        except:
            self.client.create_collection(
                collection_name=self.collection_name,
                vectors_config=VectorParams(size=384, distance=Distance.COSINE)
            )
    
    def add_document(self, file_path: str, filename: str, metadata: dict = None):
        text = self._extract_text(file_path)
        chunks = self._chunk_text(text)
        
        points = []
        for i, chunk in enumerate(chunks):
            embedding = self.encoder.encode(chunk).tolist()
            payload_metadata = {"filename": filename, "chunk_id": i}
            if metadata:
                payload_metadata.update(metadata)
            
            point = PointStruct(
                id=str(uuid.uuid4()),
                vector=embedding,
                payload={"text": chunk, "metadata": payload_metadata}
            )
            points.append(point)
        
        if points:
            self.client.upsert(collection_name=self.collection_name, points=points)
    
    def search(self, query: str, top_k: int = 5):
        query_vector = self.encoder.encode(query).tolist()
        results = self.client.search(
            collection_name=self.collection_name,
            query_vector=query_vector,
            limit=top_k
        )
        return [
            {
                "text": hit.payload["text"],
                "metadata": hit.payload["metadata"],
                "score": hit.score
            }
            for hit in results
        ]
    
    def _extract_text(self, file_path: str) -> str:
        ext = file_path.lower().split('.')[-1]
        
        if ext == 'txt':
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        elif ext == 'pdf':
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            return "\n".join([page.extract_text() for page in reader.pages])
        elif ext == 'docx':
            from docx import Document
            doc = Document(file_path)
            return "\n".join([para.text for para in doc.paragraphs])
        elif ext in ['xlsx', 'xls']:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            text = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text.append(" | ".join([str(cell) for cell in row if cell]))
            return "\n".join(text)
        else:
            return ""
    
    def _chunk_text(self, text: str, chunk_size: int = 500) -> list:
        words = text.split()
        chunks = []
        for i in range(0, len(words), chunk_size):
            chunk = " ".join(words[i:i + chunk_size])
            chunks.append(chunk)
        return chunks
    
    def is_healthy(self) -> bool:
        try:
            self.client.get_collections()
            return True
        except:
            return False
